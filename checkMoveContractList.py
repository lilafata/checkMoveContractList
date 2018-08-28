
##########################################################################################################
#
# DATE  :      08/28/2018
# AUTHOR:      Lila Fata
# FILE  :      checkMoveContractList.py
# DESCRIPTION: This script file contains the following functions:
#
#              * print_list()          - print the current available contracts in the list
#              * check_contract_date() - check if date expired for contract on top of list
#              * move_list()           - move list up when contract on top of list expires
#              * add_contract()        - add a new contract to end of the list
#
# ASSUMPTIONS: Following are assumptions to test this script -
#
#   1) The actual spreadsheet 'Problem to Solve' will only be used to analyze and manipulate
#      data and will not be overwritten in the directory where it is located
#   2) The expiration will be on the 1st (not the last day) of the month for each month listed
#   3) If contract expires, new contract 'NEW_CONTRACT' will be added to end of the list
#
# NOTES: This program was executed using the IDLE Python 3.7 Shell on a Windows 10 laptop
#
##########################################################################################################

from openpyxl import load_workbook
from datetime import datetime

CONTRACT_SHEET = './Problem to Solve.xlsx'  # Spreadsheet containing list of available contracts
NEW_CONTRACT = 'May19'                      # New contract to add to end of the list

##########################################################################################################
# Print current list of contracts available to Feeder Cattle Futures
##########################################################################################################
def print_list(sheet):
    print("\nHere's the current list of available contracts...\n")
    for i in range(1,8):
        print("value: ", sheet.cell(row=i, column=1).value)

##########################################################################################################
# Check if date expired for the contract on top of the list (expiration is 1st day of each month listed)
##########################################################################################################
def check_contract_date(sheet):
    print("\nChecking if date expired for contract on top of the list...\n")
    item = sheet.cell(row = 1, column = 1)
    dateObject = datetime.strptime(item.value, '%b%y')
    presentDate = datetime.now()
    if dateObject.date() < presentDate.date():
        print("Contract expired on", dateObject, "Today's date is:", presentDate)
        return True
    else:
        return False

##########################################################################################################
# Move list up when contract on top of list expires
##########################################################################################################
def move_list(sheet):
    print("\nMoving list up because contract on top of the list has expired...")
    for i in range(1,8):
        item = sheet.cell(row = i, column = 1)
        nextValue = sheet.cell(row = i+1, column = 1).value
        item.value = nextValue

##########################################################################################################
# Add new contract to end of list
##########################################################################################################
def add_contract(sheet, NEW_CONTRACT):
    print("\nAdding a new contract", NEW_CONTRACT, "to the end of the list...")
    for i in range(1,8):
        item = sheet.cell(row = i, column = 1)
        if not item.value:
            item.value = NEW_CONTRACT
            break
        
##########################################################################################################
# main()
##########################################################################################################
def main():
    wb = load_workbook(CONTRACT_SHEET)
    sheet = wb['Sheet1']
    print_list(sheet)                      # Print initial list of contracts
    if check_contract_date(sheet):         # Check if contract date expired
        move_list(sheet)                   # Move the list up when a contract on top of that list expires
        print_list(sheet)                  # Print current list after a contract expires
        add_contract(sheet, NEW_CONTRACT)  # Add new contract at end of list
        print_list(sheet)                  # Print current list after a new contract is added

if __name__ == "__main__":
    main()
      
